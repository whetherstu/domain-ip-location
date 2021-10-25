import time
from urllib.parse import urlparse
import openpyxl
import os
import socket
import requests
import geoip2.database
import ipv4tolocation

xlsx_set ="C:/Users/92816/Desktop/app_link.txt"
domain_set = []
def get_domain_info(domain):
    ip = ""
    lines = os.popen('nslookup ' + domain + ' 10.3.9.45')
    row = lines.readlines()
    flag = 0
    if len(row) > 4:
        for value in row[3:]:
            if 'Address: ' in value or 'Addresses: ' in value:
                ip = value[10:]
                flag = 1
        if flag == 0:
            ip = 'no'
    else:
        ip = 'no'
    return ip


# def get_ip_info(ip):
#     time.sleep(5)
#     s = requests.session()
#     requests.adapters.DEFAULT_RETRIES = 5
#     s.keep_alive = False
#     s.proxies = {"https": "47.100.104.247:8080", "http": "36.248.10.47:8080", }
#     r = requests.get('http://whois.pconline.com.cn/ipJson.jsp?ip=%s&json=true' % ip)
#     #r = requests.get('http://ip.ws.126.net/ipquery?ip=%s'%ip)
#     #r = requests.get('http://ip-api.com/json/%s' % ip)
#     # print(r.json()['status'])
#     # if r.json()['status'] == 'success':
#     i = r.json()
#     # country = i['country']  # 国家
#     #
#     # region = i['regionName']  # 地区
#     # city = i['city']  # 城市
#     # isp = i['isp']  # 运营商
#     country = i['pro']
#     region = i['addr']
#     print(region)
#         #print(u'国家: %s\n省份: %s\n城市: %s\n运营商: %s\n' % (country, region, city, isp))
#     return country,region
#
#     #      country = "ERRO! ip"
#     #      region = "ERRO! ip"
#     # #     #print("ERRO! ip: %s" % ip)
#     #      return country,region
def get_ip_info(ip):
    if "::" not in ip:
        reader = geoip2.database.Reader(
            'C:/Users/92816/Documents/WeChat Files/wxid_rl2ww9mz8mp22/FileStorage/File/2021-10/GeoLite2-City_20211019/GeoLite2-City.mmdb')
        response = reader.city(ip)
        country = response.continent.names['zh-CN']
        area = response.country.names['zh-CN']
        city = response.subdivisions.most_specific.name
        #location = ipv4tolocation.findIP(ip)#非ipv6可用，记得排除
        # country = location['country']
        # area = location['area']
        return country, area,city
    else:
        country = "no"
        area = "no"
        city = "no"
        return country, area,city



with open(xlsx_set , "rt", encoding="utf-8") as f:
    xlsx_files = f.readlines()
    f.close()

    #xlsx_file= xlsx_file[:-1]
    for xlsx_file in xlsx_files:
        xlsx_file = xlsx_file.strip('\n')
        print(xlsx_file)
        xlsx_file = eval(repr(xlsx_file).replace('\\', '/'))
        wb = openpyxl.load_workbook(xlsx_file)
        sheets = wb.sheetnames
        for sheet in sheets:
            sheet_active = wb[sheet]
            print(sheet_active)
            rows = sheet_active.max_row
            #cols = sheet_active.max_col
            for i in range(1,rows+1):
                url = sheet_active.cell(row = i,column=3).value
                domain = urlparse(url).netloc

                sheet_active.cell(row=i, column=6).value = domain
                print(domain)
                ip = get_domain_info(domain)
                #ip = socket.gethostbyname(domain)
                sheet_active.cell(row = i,column=7).value = ip
                if ip !='no':
                    print(ip)
                    country,area,city = get_ip_info(ip[:-1].strip())
                    print(country)
                    sheet_active.cell(row=i, column=8).value = country
                    sheet_active.cell(row=i, column=9).value = area
                    sheet_active.cell(row=i, column=10).value = city
            wb.save(xlsx_file)
        wb.save(xlsx_file)
        time.sleep(2)